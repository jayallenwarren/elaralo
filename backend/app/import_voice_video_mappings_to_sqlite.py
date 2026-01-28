#!/usr/bin/env python3
"""Import the 'Voice and Video Mappings - Elaralo.xlsx' spreadsheet into a SQLite DB.

Usage:
  python import_voice_video_mappings_to_sqlite.py \
    --xlsx "Voice and Video Mappings - Elaralo.xlsx" \
    --out "companion_catalog.sqlite"

This script is intended for *offline* use (build-time / local dev) so the API runtime does not need
openpyxl/pandas. The API only needs the resulting companion_catalog.sqlite file.
"""

from __future__ import annotations

import argparse
import os
import sqlite3
from typing import Any, Optional

import openpyxl


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to the Excel file")
    ap.add_argument("--out", required=True, help="Path to write the sqlite DB")
    args = ap.parse_args()

    xlsx_path = args.xlsx
    out_path = args.out

    if not os.path.exists(xlsx_path):
        raise SystemExit(f"XLSX not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    # Default to the first sheet
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    header_to_idx = {str(h).strip(): i for i, h in enumerate(headers)}

    required = [
        "Brand",
        "Avatar",
        "Communication",
        "Eleven Labs Voice ID",
        "Live",
        "D-ID Agent ID",
        "D-ID Client Key",
        "D-ID Agent Link",
        "D-ID Embed Code",
        "ElevenLabs Voice",
    ]
    missing = [h for h in required if h not in header_to_idx]
    if missing:
        raise SystemExit(f"Missing expected columns: {missing}")

    if os.path.exists(out_path):
        os.remove(out_path)

    conn = sqlite3.connect(out_path)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE companion_mappings (
          brand TEXT NOT NULL,
          avatar TEXT NOT NULL,
          eleven_voice_name TEXT,
          communication TEXT NOT NULL,
          eleven_voice_id TEXT,
          live_provider TEXT,
          did_embed_code TEXT,
          did_agent_link TEXT,
          did_agent_id TEXT,
          did_client_key TEXT,
          PRIMARY KEY (brand, avatar)
        );
        """
    )

    inserted = 0
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in row):
            continue

        def col(name: str) -> Optional[str]:
            return _clean(row[header_to_idx[name]])

        brand = col("Brand")
        avatar = col("Avatar")
        if not brand or not avatar:
            continue

        cur.execute(
            """
            INSERT OR REPLACE INTO companion_mappings
            (brand, avatar, eleven_voice_name, communication, eleven_voice_id, live_provider, did_embed_code, did_agent_link, did_agent_id, did_client_key)
            VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (
                brand,
                avatar,
                col("ElevenLabs Voice"),
                col("Communication") or "Audio",
                col("Eleven Labs Voice ID"),
                col("Live"),
                col("D-ID Embed Code"),
                col("D-ID Agent Link"),
                col("D-ID Agent ID"),
                col("D-ID Client Key"),
            ),
        )
        inserted += 1

    conn.commit()
    conn.close()

    print(f"Wrote {inserted} rows to {out_path}")


if __name__ == "__main__":
    main()
